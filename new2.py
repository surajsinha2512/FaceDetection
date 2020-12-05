def excel():
    import os
    cwd = os.getcwd()
    os.chdir("C:/Users/ANIKET SINHA/Desktop/")   
    import xlsxwriter 
    workbook = xlsxwriter.Workbook('example.xlsx') 
    worksheet = workbook.add_worksheet() 
    worksheet.write('A1', 'Name') 
    worksheet.write('B1', 'Roll') 
    worksheet.write('C1', 'Branch') 
    worksheet.write('D1', 'Attendance') 
    worksheet.write('A2', 'Suraj') 
    worksheet.write('B2', 9) 
    worksheet.write('C2', 'IT') 
    worksheet.write('D2', i) 
    worksheet.write('A3', 'Rajeev') 
    worksheet.write('B3', 26) 
    worksheet.write('C3', 'IT') 
    worksheet.write('D3', j) 
    worksheet.write('A4', 'Rajdeep') 
    worksheet.write('B4', 38) 
    worksheet.write('C4', 'IT') 
    worksheet.write('D4', k) 
    worksheet.write('A5', 'Sherap') 
    worksheet.write('B5', 54) 
    worksheet.write('C5', 'IT') 
    worksheet.write('D5', k) 
    k=input("enter coordinate")
    j=input("enter roll")
    worksheet.write(k, j) 
    workbook.close() 

def excel():
    import openpyxl
	import os
    cwd = os.getcwd()
    os.chdir("C:/Users/ANIKET SINHA/Desktop/")  
    wbkName = 'example.xlsx'
    wbk = openpyxl.load_workbook(wbkName)
    wks=wbk.active
    wks.cell(row, column).value =1
    wbk.save(wbkName)
    wbk.close